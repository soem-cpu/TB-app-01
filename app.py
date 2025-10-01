import streamlit as st
import pandas as pd
import importlib.util
import io

# =========================
# Page Config
# =========================
st.set_page_config(
    page_title="TB Data Verification",
    layout="wide",
    page_icon="📊"
)

# =========================
# Sidebar
# =========================
with st.sidebar:
    st.image(
        "https://www.who.int/images/default-source/imported/tuberculosis.jpg", 
        caption="Stop TB Initiative", 
        use_container_width=True
    )
    st.markdown("### 📌 Instructions")
    st.markdown("""
    1. Upload **Rules file (.py)**
    2. Upload **Excel/CSV data**
    3. Review validation results
    4. Download corrected file
    """)
    st.markdown("---")
    st.info("💡 Make sure the dataset column names match the rules file.")

# =========================
# Title & Description
# =========================
st.title("🩺 Tuberculosis Data Verification System")
st.markdown("""
Upload your **Python rules file** and a **dataset (Excel/CSV)**.  
This tool applies dynamic validation rules and returns a downloadable Excel report.
""")

# =========================
# Upload Rules File
# =========================
st.subheader("📥 Step 1: Upload Rules File")
rules_file = st.file_uploader("Upload Python rules file (.py)", type=["py"])
rules_module = None

if rules_file:
    with open("rules_temp.py", "wb") as f:
        f.write(rules_file.getbuffer())
    spec = importlib.util.spec_from_file_location("rules_module", "rules_temp.py")
    rules_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(rules_module)
    st.success("✅ Rules file loaded successfully!")

import openpyxl

if data_file.name.endswith("xlsx"):
    try:
        wb = openpyxl.load_workbook(data_file)
        preview_sheet = wb.sheetnames[0]
        ws = wb[preview_sheet]

        # Read first 10 rows manually into dataframe
        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=10, values_only=True)]
        df_preview = pd.DataFrame(data[1:], columns=data[0])  # first row as header

        st.write(f"Preview of uploaded data (first sheet: {preview_sheet}):")
        st.dataframe(df_preview.head())
    except Exception as e:
        st.error(f"Could not preview Excel file: {e}")




# =========================
# Upload Data File
# =========================
st.subheader("📥 Step 2: Upload Data File")
data_file = st.file_uploader("Upload Excel file to verify", type=["xlsx", "csv"])
if data_file and rules_file:
    # Preview logic
    if data_file.name.endswith("xlsx"):
        try:
            xls = pd.ExcelFile(data_file, engine="openpyxl")  # 👈 force openpyxl
            preview_sheet = xls.sheet_names[0]
            df_preview = xls.parse(preview_sheet, engine="openpyxl")
            st.write(f"Preview of uploaded data (first sheet: {preview_sheet}):")
            st.dataframe(df_preview.head())
        except Exception as e:
            st.error(f"Could not preview Excel file: {e}")

# =========================
# Run Rules if Both Uploaded
# =========================
if data_file and rules_module:
    st.subheader("👀 Step 3: Data Preview")

    try:
        if data_file.name.endswith("xlsx"):
            xls = pd.ExcelFile(data_file)
            preview_sheet = xls.sheet_names[0]
            df_preview = xls.parse(preview_sheet)
            st.info(f"Preview (first sheet: **{preview_sheet}**)")
        else:
            df_preview = pd.read_csv(data_file)
            st.info("Preview (CSV)")
        st.dataframe(df_preview.head())
    except Exception as e:
        st.error(f"❌ Could not preview data: {e}")

    st.subheader("🔍 Step 4: Run Rule Checks")
    try:
        results = rules_module.check_rules(data_file)
        excel_output = io.BytesIO()
        sheet_count = 0

        with pd.ExcelWriter(excel_output, engine="xlsxwriter") as writer:
            if isinstance(results, dict):
                st.markdown("## 📑 Validation Results")
                for k, v in results.items():
                    st.markdown(f"### 📂 {k}")
                    if isinstance(v, pd.DataFrame):
                        if not v.empty:
                            st.dataframe(v, use_container_width=True)
                        else:
                            st.success(f"No issues in **{k}** ✅")
                        v.to_excel(writer, index=False, sheet_name=k[:31])
                        sheet_count += 1
            elif isinstance(results, pd.DataFrame):
                if results.empty:
                    st.success("✅ No validation issues found!")
                else:
                    st.dataframe(results, use_container_width=True)
                results.to_excel(writer, index=False, sheet_name="Validation")
                sheet_count += 1

        if sheet_count > 0:
            st.download_button(
                label="⬇️ Download Validation Report (Excel)",
                data=excel_output.getvalue(),
                file_name="TB_Validation_Results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"❌ Error running rules: {e}")

# =========================
# Footer
# =========================
st.markdown("---")
st.caption("🔬 TB Data Quality Assurance | Powered by Streamlit")
